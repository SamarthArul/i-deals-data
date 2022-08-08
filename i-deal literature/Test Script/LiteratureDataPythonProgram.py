'''
Created on Jun 21, 2021

@author: samartharul
'''
print("openpyxl test for idiosyncratic deals literature gathering")

import openpyxl
from openpyxl import Workbook, load_workbook


wb = load_workbook('I-DealLiterature.xlsx')
ws = wb.active
print(ws['A2'].value)
print(ws['A2'].value)
print(ws['A3'].value)
print(ws['A4'].value)
print(ws['A5'].value)
print(ws['A6'].value)
print(ws['A7'].value)
print(ws['A8'].value)
print(ws['A9'].value)
print(ws['A10'].value)
print(ws['A11'].value)
print(ws['A12'].value)


print(ws['B2'].value)
print(ws['B2'].value)
print(ws['B3'].value)
print(ws['B4'].value)
print(ws['B5'].value)
print(ws['B6'].value)
print(ws['B7'].value)
print(ws['B8'].value)
print(ws['B9'].value)
print(ws['B10'].value)
print(ws['B11'].value)
print(ws['B12'].value)
